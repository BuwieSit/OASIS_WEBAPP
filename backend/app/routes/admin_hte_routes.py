from flask import Blueprint, jsonify, request, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt
from sqlalchemy.orm import aliased
from datetime import date, datetime, timedelta
import io
import os
import uuid
import requests
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import HostTrainingEstablishment, MemorandumOfAgreement
from app.models.user import UserRole

admin_hte_bp = Blueprint("admin_hte_bp", __name__, url_prefix="/api/admin/htes")

EXCEL_HEADERS = [
    "COMPANY NAME", "NATURE OF BUSINESS", "CONTACT PERSON", "POSITION",
    "CONTACT NUMBER", "EMAIL ADDRESS", "COMPANY ADDRESS", "MOA STATUS",
    "COURSE", "DATE NOTARIZED", "VALIDITY", "EXPIRY DATE",
    "LINKE TO SCANNED MOA",
]


def _admin_only():
    claims = get_jwt()
    return claims.get("role") == UserRole.ADMIN.value


def _parse_date(val):
    if val is None or val == "":
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    if isinstance(val, (int, float)):
        # Year-only values like 2022
        if 1900 <= int(val) <= 2100:
            return date(int(val), 1, 1)

        # Excel serial date values like 44662
        try:
            return date(1899, 12, 30) + timedelta(days=int(val))
        except Exception:
            return None

    s = str(val).strip()

    if s.isdigit() and 1900 <= int(s) <= 2100:
        return date(int(s), 1, 1)

    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    return None


def _parse_number(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except Exception:
        return None


def _parse_validity_months(val):
    if val is None or val == "":
        return None

    if isinstance(val, (int, float)):
        return int(float(val))

    s = str(val).strip().upper()

    if not s:
        return None

    number = "".join(ch for ch in s if ch.isdigit() or ch == ".")

    if "YEAR" in s:
        return int(float(number) * 12) if number else None

    if "MONTH" in s:
        return int(float(number)) if number else None

    try:
        return int(float(s))
    except Exception:
        return None


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_moa_status(value):
    status = _normalize_text(value).upper()

    if status in {"WITH MOA", "WITHMOA", "ACTIVE", "APPROVED"}:
        return "ACTIVE"

    if status in {"EXPIRED"}:
        return "EXPIRED"

    return "PENDING"


def _compute_validity_years(signed_at, expires_at):
    if not signed_at or not expires_at:
        return None

    delta_days = (expires_at - signed_at).days
    if delta_days < 0:
        return None

    return round(delta_days / 365, 2)


def _compute_validity_years_from_months(validity_months):
    if validity_months is None:
        return None

    try:
        validity_months = float(validity_months)
    except Exception:
        return None

    if validity_months < 0:
        return None

    return round(validity_months / 12, 2)


def _compute_expiry_date_from_months(signed_at, validity_months):
    if not signed_at or validity_months is None:
        return None

    try:
        validity_months = int(float(validity_months))
    except Exception:
        return None

    if validity_months < 0:
        return None

    year = signed_at.year
    month = signed_at.month + validity_months
    year += (month - 1) // 12
    month = ((month - 1) % 12) + 1

    day = min(signed_at.day, 28)
    return date(year, month, day)


def _resolve_moa_dates_and_validity(signed_at=None, expires_at=None, validity_months=None, validity_years=None):
    if signed_at and validity_months is not None:
        expires_at = _compute_expiry_date_from_months(signed_at, validity_months)
        validity_years = _compute_validity_years_from_months(validity_months)
    elif signed_at and expires_at and validity_years is None:
        validity_years = _compute_validity_years(signed_at, expires_at)

    return signed_at, expires_at, validity_years


def _save_upload(file_storage, upload_root, subdir, allowed_ext=None):
    if not file_storage:
        return None

    filename = secure_filename(file_storage.filename or "")
    if not filename:
        return None

    ext = os.path.splitext(filename)[1].lower()
    if allowed_ext and ext not in allowed_ext:
        raise ValueError(f"Invalid file type: {ext}")

    out_dir = os.path.join(upload_root, subdir)
    os.makedirs(out_dir, exist_ok=True)

    base = secure_filename(os.path.splitext(filename)[0])
    out_name = f"{base}_{uuid.uuid4().hex}{ext}"
    out_path = os.path.join(out_dir, out_name)
    file_storage.save(out_path)

    return f"uploads/{subdir}/{out_name}"


def _overview_query(status):
    latest_moa = (
        db.session.query(
            MemorandumOfAgreement.hte_id,
            db.func.max(MemorandumOfAgreement.created_at).label("latest_created"),
        )
        .group_by(MemorandumOfAgreement.hte_id)
        .subquery()
    )

    moa = aliased(MemorandumOfAgreement)

    query = (
        db.session.query(HostTrainingEstablishment, moa)
        .outerjoin(latest_moa, latest_moa.c.hte_id == HostTrainingEstablishment.id)
        .outerjoin(
            moa,
            (moa.hte_id == HostTrainingEstablishment.id)
            & (moa.created_at == latest_moa.c.latest_created)
        )
    )

    if status and status.upper() != "ALL":
        query = query.filter(moa.status == status.upper())

    return query.order_by(HostTrainingEstablishment.company_name)


def _clean_moa_link(value):
    link = _normalize_text(value)
    if not link:
        return ""

    lowered = link.lower()
    invalid_values = {
        "link to scanned moa",
        "linke to scanned moa",
        "n/a",
        "na",
        "-",
        "--",
        "none",
        "null",
    }

    if lowered in invalid_values:
        return ""

    return link


def _is_http_url(value):
    if not value:
        return False
    value = str(value).strip().lower()
    return value.startswith("http://") or value.startswith("https://")


def download_gdrive_file_to_bytes(url):
    def _get_confirm_token(response):
        for key, value in response.cookies.items():
            if key.startswith("download_warning"):
                return value
        return None

    try:
        if not url:
            print("GDRIVE: empty URL")
            return None

        url = str(url).strip()

        if not _is_http_url(url):
            print("GDRIVE: invalid URL format:", url)
            return None

        if "drive.google.com" not in url:
            print("GDRIVE: non-drive URL, skipping:", url)
            return None

        if "id=" in url:
            file_id = url.split("id=")[-1].split("&")[0]
        elif "/d/" in url:
            file_id = url.split("/d/")[1].split("/")[0]
        else:
            print("GDRIVE: could not extract file id from URL:", url)
            return None

        session = requests.Session()
        base_url = "https://drive.google.com/uc?export=download"

        response = session.get(base_url, params={"id": file_id}, stream=True, timeout=20)

        token = _get_confirm_token(response)
        if token:
            response = session.get(
                base_url,
                params={"id": file_id, "confirm": token},
                stream=True,
                timeout=20,
            )

        if response.status_code != 200:
            print(f"GDRIVE: bad status {response.status_code} for {url}")
            return None

        content_type = (response.headers.get("Content-Type") or "").lower()

        if "text/html" in content_type:
            print(f"GDRIVE: got HTML instead of file for {url}")
            return None

        filename = f"moa_{uuid.uuid4().hex}.pdf"
        content_disp = response.headers.get("Content-Disposition") or ""

        if "filename=" in content_disp:
            raw_name = content_disp.split("filename=")[-1].strip().strip('"')
            if raw_name:
                filename = secure_filename(raw_name)

        file_bytes = response.content
        if not file_bytes:
            print(f"GDRIVE: empty file bytes for {url}")
            return None

        return {
            "filename": filename,
            "mime_type": response.headers.get("Content-Type", "application/pdf"),
            "blob": file_bytes,
            "size": len(file_bytes),
        }

    except Exception as e:
        print("GDRIVE download failed:", e)
        return None


@admin_hte_bp.get("")
@jwt_required()
def get_htes():
    if not _admin_only():
        return jsonify({"error": "forbidden"}), 403

    status = request.args.get("status")
    rows = _overview_query(status).all()

    results = []
    for hte, moa in rows:
        signed_at = moa.signed_at if moa and moa.signed_at else hte.moa_signed_at
        expires_at = moa.expires_at if moa and moa.expires_at else hte.moa_expiry_date
        validity_years = hte.moa_validity

        if signed_at and expires_at and validity_years is None:
            validity_years = _compute_validity_years(signed_at, expires_at)

        results.append({
            "hte_id": hte.id,
            "company_name": hte.company_name,
            "industry": hte.industry,
            "address": hte.address,
            "description": hte.description,
            "website": hte.website,
            "contact_person": hte.contact_person,
            "contact_position": hte.contact_position,
            "contact_number": hte.contact_number,
            "contact_email": hte.contact_email,
            "course": hte.course,
            "thumbnail_path": hte.thumbnail_path,
            "moa": None if not moa else {
                "id": moa.id,
                "status": moa.status,
                "signed_at": signed_at.isoformat() if signed_at else None,
                "expires_at": expires_at.isoformat() if expires_at else None,
                "validity_years": validity_years,
                "document_path": moa.document_path,
                "document_filename": getattr(moa, "document_filename", None),
                "document_size": getattr(moa, "document_size", None),
            },
            "moa_validity_years": validity_years
        })

    return jsonify(results), 200


@admin_hte_bp.get("/export")
@jwt_required()
def export_htes_excel():
    if not _admin_only():
        return jsonify({"error": "forbidden"}), 403

    status = request.args.get("status")
    rows = _overview_query(status).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "HTE Overview"
    ws.append(EXCEL_HEADERS)

    for hte, moa in rows:
        signed_at = moa.signed_at if moa and moa.signed_at else hte.moa_signed_at
        expires_at = moa.expires_at if moa and moa.expires_at else hte.moa_expiry_date
        validity_years = hte.moa_validity

        if signed_at and expires_at and validity_years is None:
            validity_years = _compute_validity_years(signed_at, expires_at)

        ws.append([
            hte.company_name,
            hte.industry,
            hte.contact_person,
            hte.contact_position,
            hte.contact_number,
            hte.contact_email,
            hte.address,
            (moa.status if moa else hte.moa_status),
            hte.course or "",
            (signed_at.strftime("%m/%d/%Y") if signed_at else ""),
            (validity_years if validity_years is not None else ""),
            (expires_at.strftime("%m/%d/%Y") if expires_at else ""),
            (moa.document_path if moa else hte.moa_file_path) or "",
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"hte_overview_{status or 'ALL'}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_hte_bp.post("/import")
@jwt_required()
def import_htes_excel():
    if not _admin_only():
        return jsonify({"error": "forbidden"}), 403

    if "file" not in request.files:
        return jsonify({"error": "file_required"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "invalid_file_type", "message": "Only .xlsx allowed"}), 400

    wb = load_workbook(f, data_only=True)
    ws = wb.active

    header_row = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    header_to_idx = {h: i for i, h in enumerate(header_row)}

    missing = [h for h in EXCEL_HEADERS if h not in header_to_idx]
    if missing:
        return jsonify({
            "error": "invalid_template",
            "message": f"Missing headers: {', '.join(missing)}"
        }), 400

    created_htes = 0
    updated_htes = 0
    created_moas = 0
    updated_moas = 0
    failed_rows = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def get(h):
            return row[header_to_idx[h]].value

        company_name = _normalize_text(get("COMPANY NAME"))

        if not company_name or company_name.upper() == "COMPANY NAME":
            failed_rows.append({"row": r_idx, "error": "Missing COMPANY NAME"})
            continue

        industry = _normalize_text(get("NATURE OF BUSINESS")) or "N/A"
        contact_person = _normalize_text(get("CONTACT PERSON")) or "N/A"
        contact_position = _normalize_text(get("POSITION")) or "N/A"
        contact_number = _normalize_text(get("CONTACT NUMBER")) or "N/A"
        contact_email = _normalize_text(get("EMAIL ADDRESS")) or "N/A"
        address = _normalize_text(get("COMPANY ADDRESS")) or "N/A"
        moa_status = _normalize_moa_status(get("MOA STATUS"))
        course = _normalize_text(get("COURSE")) or None
        signed_at = _parse_date(get("DATE NOTARIZED"))
        validity_months = _parse_validity_months(get("VALIDITY"))
        expires_at = _parse_date(get("EXPIRY DATE"))
        moa_link = _clean_moa_link(get("LINKE TO SCANNED MOA"))

        moa_file_data = None

        if moa_link:
            if _is_http_url(moa_link):
                moa_file_data = download_gdrive_file_to_bytes(moa_link)

                if not moa_file_data:
                    failed_rows.append({
                        "row": r_idx,
                        "error": "MOA link could not be downloaded",
                        "link": moa_link
                    })
            else:
                failed_rows.append({
                    "row": r_idx,
                    "error": "Invalid MOA link format",
                    "link": moa_link
                })

        signed_at, expires_at, validity_years = _resolve_moa_dates_and_validity(
            signed_at=signed_at,
            expires_at=expires_at,
            validity_months=validity_months,
            validity_years=None
        )

        try:
            with db.session.begin_nested():
                hte = HostTrainingEstablishment.query.filter_by(company_name=company_name).first()

                if not hte:
                    hte = HostTrainingEstablishment(
                        company_name=company_name,
                        industry=industry,
                        address=address,
                        description=None,
                        website=None,
                        contact_person=contact_person,
                        contact_position=contact_position,
                        contact_number=contact_number,
                        contact_email=contact_email,
                        moa_status=moa_status,
                        course=course,
                        moa_signed_at=signed_at,
                        moa_validity=validity_years,
                        moa_expiry_date=expires_at,
                        moa_file_path=moa_link if moa_link else None,
                    )
                    db.session.add(hte)
                    db.session.flush()
                    created_htes += 1
                else:
                    hte.industry = industry or hte.industry
                    hte.address = address or hte.address
                    hte.contact_person = contact_person or hte.contact_person
                    hte.contact_position = contact_position or hte.contact_position
                    hte.contact_number = contact_number or hte.contact_number
                    hte.contact_email = contact_email or hte.contact_email
                    hte.course = course or hte.course
                    hte.moa_status = moa_status
                    hte.moa_signed_at = signed_at
                    hte.moa_validity = validity_years
                    hte.moa_expiry_date = expires_at
                    if moa_link:
                        hte.moa_file_path = moa_link
                    updated_htes += 1

                if signed_at and expires_at:
                    existing_moa = (
                        MemorandumOfAgreement.query
                        .filter_by(hte_id=hte.id, signed_at=signed_at, expires_at=expires_at)
                        .first()
                    )

                    if not existing_moa:
                        db.session.add(MemorandumOfAgreement(
                            hte_id=hte.id,
                            signed_at=signed_at,
                            expires_at=expires_at,
                            status=moa_status,
                            document_path=moa_link if moa_link else None,
                            document_filename=moa_file_data["filename"] if moa_file_data else None,
                            document_mime_type=moa_file_data["mime_type"] if moa_file_data else None,
                            document_blob=moa_file_data["blob"] if moa_file_data else None,
                            document_size=moa_file_data["size"] if moa_file_data else None,
                        ))
                        created_moas += 1
                    else:
                        existing_moa.status = moa_status
                        if moa_link:
                            existing_moa.document_path = moa_link

                        if moa_file_data:
                            existing_moa.document_filename = moa_file_data["filename"]
                            existing_moa.document_mime_type = moa_file_data["mime_type"]
                            existing_moa.document_blob = moa_file_data["blob"]
                            existing_moa.document_size = moa_file_data["size"]

                        updated_moas += 1
                else:
                    failed_rows.append({
                        "row": r_idx,
                        "error": "MOA not created because DATE NOTARIZED and/or VALIDITY/EXPIRY DATE could not be parsed"
                    })

        except Exception as e:
            failed_rows.append({
                "row": r_idx,
                "error": f"Row processing failed: {str(e)}"
            })

    db.session.commit()

    return jsonify({
        "created_htes": created_htes,
        "updated_htes": updated_htes,
        "created_moas": created_moas,
        "updated_moas": updated_moas,
        "failed_rows": failed_rows
    }), 200


@admin_hte_bp.post("")
@jwt_required()
def create_hte_manual():
    claims = get_jwt()
    if claims.get("role") != UserRole.ADMIN.value:
        return jsonify({"error": "forbidden"}), 403

    upload_root = current_app.config.get("UPLOAD_ROOT")
    if not upload_root:
        return jsonify({"error": "server_config", "message": "UPLOAD_ROOT not set"}), 500

    form = request.form

    company_name = (form.get("company_name") or "").strip()
    industry = (form.get("industry") or "").strip()
    address = (form.get("address") or "").strip()

    contact_person = (form.get("contact_person") or "").strip()
    contact_position = (form.get("contact_position") or "").strip()
    contact_number = (form.get("contact_number") or "").strip()
    contact_email = (form.get("contact_email") or "").strip()

    if not company_name or not industry or not address:
        return jsonify({"error": "validation_error", "message": "company_name, industry, address are required"}), 400

    if not contact_person or not contact_position or not contact_number or not contact_email:
        return jsonify({"error": "validation_error", "message": "Contact fields are required"}), 400

    description = (form.get("description") or "").strip() or None
    website = (form.get("website") or "").strip() or None

    eligible_courses_raw = form.get("eligible_courses")
    course_value = eligible_courses_raw if eligible_courses_raw else None

    status = _normalize_moa_status(form.get("status"))

    signed_at = _parse_date(form.get("signed_at"))
    validity_months = _parse_validity_months(form.get("validity"))

    expires_at = None
    validity_years = None

    signed_at, expires_at, validity_years = _resolve_moa_dates_and_validity(
        signed_at=signed_at,
        expires_at=expires_at,
        validity_months=validity_months,
        validity_years=validity_years
    )

    logo = request.files.get("logo")
    thumbnail = request.files.get("thumbnail")
    moa_file = request.files.get("moa_file")

    try:
        logo_path = _save_upload(
            logo, upload_root, "hte_logos",
            allowed_ext={".png", ".jpg", ".jpeg", ".webp"}
        ) if logo else None

        thumbnail_path = _save_upload(
            thumbnail, upload_root, "hte_thumbnails",
            allowed_ext={".png", ".jpg", ".jpeg", ".webp"}
        ) if thumbnail else None

    except ValueError as e:
        return jsonify({"error": "validation_error", "message": str(e)}), 400

    moa_file_data = None
    if moa_file:
        filename = secure_filename(moa_file.filename or "") or f"moa_{uuid.uuid4().hex}.pdf"
        mime_type = moa_file.mimetype or "application/pdf"
        blob = moa_file.read()

        if not blob:
            return jsonify({"error": "validation_error", "message": "Uploaded MOA file is empty"}), 400

        moa_file_data = {
            "filename": filename,
            "mime_type": mime_type,
            "blob": blob,
            "size": len(blob),
        }

    hte = HostTrainingEstablishment(
        company_name=company_name,
        industry=industry,
        address=address,
        description=description,
        website=website,
        contact_person=contact_person,
        contact_position=contact_position,
        contact_number=contact_number,
        contact_email=contact_email,
        moa_status=status,
        course=course_value,
        moa_signed_at=signed_at,
        moa_validity=validity_years,
        moa_expiry_date=expires_at,
        moa_file_path=None,
        thumbnail_path=thumbnail_path,
    )

    if hasattr(hte, "logo_path"):
        hte.logo_path = logo_path

    db.session.add(hte)
    db.session.flush()

    if signed_at and expires_at:
        moa = MemorandumOfAgreement(
            hte_id=hte.id,
            signed_at=signed_at,
            expires_at=expires_at,
            status=status,
            document_path=None,
            document_filename=moa_file_data["filename"] if moa_file_data else None,
            document_mime_type=moa_file_data["mime_type"] if moa_file_data else None,
            document_blob=moa_file_data["blob"] if moa_file_data else None,
            document_size=moa_file_data["size"] if moa_file_data else None,
        )
        db.session.add(moa)

    db.session.commit()

    return jsonify({
        "message": "created",
        "hte_id": hte.id,
        "signed_at": signed_at.isoformat() if signed_at else None,
        "validity_months": validity_months,
        "validity_years": validity_years,
        "expires_at": expires_at.isoformat() if expires_at else None
    }), 201